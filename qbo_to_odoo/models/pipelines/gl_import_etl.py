"""QuickBooks Online Journal Export Import Pipeline

Imports transactions from a QBO Journal XLSX export that are not
available via the QBO API (primarily Payroll Cheques).

The Journal export contains every debit and credit in QBO, grouped by
transaction ID.  Each transaction has the same ID used by the QBO API,
so we can skip any transaction already imported via the API pipelines
and only import the ones that are missing.

The field on ``qbo.connection`` is called ``gl_export_file`` for
backwards compatibility but should be populated with the Journal
export (not the General Ledger export).
"""

import base64
import logging
from collections import defaultdict
from io import BytesIO
from typing import Dict, List, Set, Tuple

from odoo import models

from odoo.addons.etl_framework import ETL, ETLContext, post_lock

from .extractor import QBOExtractor

_logger = logging.getLogger(__name__)

# QBO API field name → account.move field mapping
_QBO_ID_FIELDS = [
    "qbo_invoice_id",
    "qbo_bill_id",
    "qbo_expense_id",
    "qbo_transfer_id",
    "qbo_deposit_id",
    "qbo_journal_entry_id",
    "qbo_credit_memo_id",
    "qbo_vendor_credit_id",
    "qbo_sales_receipt_id",
    "qbo_refund_receipt_id",
    "qbo_tax_payment_id",
    "qbo_cc_payment_id",
]

# account.payment QBO fields
_QBO_PAYMENT_FIELDS = [
    "qbo_payment_id",
    "qbo_bill_payment_id",
]


def _parse_journal_export(file_content: bytes) -> List[Dict]:
    """Parse a QBO Journal XLSX export into grouped transactions.

    Returns a list of transactions, each with an ``id``, ``type``,
    ``date``, ``num``, ``name``, and ``lines`` (list of debit/credit
    dicts with account_code, account_name, debit, credit, memo).
    """
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
    ws = wb.active

    transactions = []
    current_id = None
    current_lines = []
    has_txn_id_col = False

    for row in ws.iter_rows(min_row=5, max_row=5, values_only=True):
        # Detect Transaction ID column (index 10)
        if len(row) > 10 and row[10] and "Transaction ID" in str(row[10]):
            has_txn_id_col = True
        break

    for row in ws.iter_rows(min_row=6, values_only=True):
        col0 = row[0]
        txn_date = row[1]
        txn_type = row[2]
        txn_num = row[3]
        name = row[4]
        memo = row[5]
        acct_code = row[6]
        acct_name = row[7]
        debit = row[8]
        credit = row[9]
        txn_id_col = row[10] if len(row) > 10 else None

        # Transaction header (ID)
        if col0 and not txn_date:
            col0_str = str(col0).strip()
            if col0_str.startswith("Total for"):
                # End of transaction — save it
                if current_id and current_lines:
                    transactions.append({
                        "id": current_id,
                        "lines": current_lines,
                    })
                current_id = None
                current_lines = []
            elif col0_str.startswith("TOTAL"):
                continue
            else:
                current_id = col0_str
                current_lines = []
            continue

        # Transaction line — use Transaction ID column if available,
        # otherwise fall back to the group header ID
        if txn_type and acct_code is not None:
            if has_txn_id_col and txn_id_col:
                line_txn_id = str(int(txn_id_col)) if isinstance(txn_id_col, float) else str(txn_id_col)
                # Update current_id from the column for accuracy
                current_id = line_txn_id
            # Parse date (DD/MM/YYYY → YYYY-MM-DD)
            date_str = str(txn_date) if txn_date else None
            if date_str and "/" in date_str:
                parts = date_str.split("/")
                if len(parts) == 3:
                    date_str = f"{parts[2]}-{parts[1]}-{parts[0]}"

            try:
                d = float(debit) if debit else 0.0
            except (ValueError, TypeError):
                d = 0.0
            try:
                c = float(credit) if credit else 0.0
            except (ValueError, TypeError):
                c = 0.0

            if d == 0 and c == 0:
                continue

            current_lines.append({
                "date": date_str,
                "type": str(txn_type),
                "num": str(txn_num or ""),
                "name": str(name or ""),
                "memo": str(memo or ""),
                "account_code": str(acct_code).strip(),
                "account_name": str(acct_name or ""),
                "debit": d,
                "credit": c,
            })

    wb.close()
    return transactions


def _get_imported_qbo_ids(ctx) -> Set[str]:
    """Collect all QBO transaction IDs already in Odoo."""
    imported = set()
    cr = ctx.env.cr

    for field in _QBO_ID_FIELDS:
        cr.execute(
            f"SELECT {field} FROM account_move "
            f"WHERE {field} IS NOT NULL AND {field} != 0"
        )
        imported.update(str(row[0]) for row in cr.fetchall())

    for field in _QBO_PAYMENT_FIELDS:
        cr.execute(
            f"SELECT {field} FROM account_payment "
            f"WHERE {field} IS NOT NULL AND {field} != 0"
        )
        imported.update(str(row[0]) for row in cr.fetchall())

    return imported


@ETL.pipeline(
    target_model="account.move",
    importer_name="qbo.gl.import",
    sap_source="GLExport",
    depends_on=[
        "qbo.account.importer",
        "qbo.tax.importer",
        "qbo.payment.importer",
    ],
)
class QboJournalImporter(models.AbstractModel):
    """ETL Pipeline for importing transactions from a QBO Journal export.

    Imports only transactions whose QBO ID is not already in Odoo,
    ensuring no double-dipping with the API-based pipelines.
    """

    _name = "qbo.gl.import"
    _description = "QBO Journal Export Importer"

    @ETL.extract("GLExport")
    def extract_journal_transactions(self, ctx: ETLContext) -> List[Dict]:
        """Parse Journal export and filter to unimported transactions."""
        extractor = QBOExtractor(ctx)

        connection = ctx.env["qbo.connection"].browse(
            ctx.get_config("source_id")
        )
        if not connection.gl_export_file:
            _logger.info(
                "No Journal export file uploaded — skipping Journal import"
            )
            return {"transactions": [], "extractor": extractor.export()}

        file_content = base64.b64decode(connection.gl_export_file)
        transactions = _parse_journal_export(file_content)
        _logger.info(f"Parsed {len(transactions)} transactions from Journal export")

        # Get all already-imported QBO IDs
        imported_ids = _get_imported_qbo_ids(ctx)
        _logger.info(f"Found {len(imported_ids)} already-imported QBO IDs")

        # Filter to only unimported transactions
        new_txns = [
            t for t in transactions if str(t["id"]) not in imported_ids
        ]

        # Count by type
        type_counts = defaultdict(int)
        for t in new_txns:
            if t["lines"]:
                type_counts[t["lines"][0]["type"]] += 1

        _logger.info(
            f"After filtering: {len(new_txns)} unimported transactions "
            f"({', '.join(f'{t}: {c}' for t, c in sorted(type_counts.items()))})"
        )

        extractor.preload("account")
        extractor.preload_journals("general")

        # Build code → account_id map and account → currency map
        ctx.env.cr.execute("""
            SELECT aa.id,
                   aa.code_store::jsonb->>jsonb_object_keys(
                       aa.code_store::jsonb) as code,
                   aa.currency_id
            FROM account_account aa
        """)
        code_map = {}
        account_currency_map = {}
        for acct_id, code, currency_id in ctx.env.cr.fetchall():
            if code:
                code_map[code] = acct_id
                if currency_id:
                    account_currency_map[acct_id] = currency_id
        extractor.extra["code_map"] = code_map
        extractor.extra["account_currency_map"] = account_currency_map

        # Company currency ID
        company_currency_id = ctx.env.company.currency_id.id
        extractor.extra["company_currency_id"] = company_currency_id

        return {"transactions": new_txns, "extractor": extractor.export()}

    @ETL.transform()
    def transform_journal_transactions(
        self, ctx: ETLContext, extracted: Dict
    ) -> List[Dict]:
        """Transform Journal transactions into account.move values."""
        data = extracted.get("extract_journal_transactions", {})
        transactions = data.get("transactions", [])
        extractor_data = data.get("extractor", {})

        if not transactions:
            return []

        from .move_builder import QBOMoveBuilder

        builder = QBOMoveBuilder(extractor_data)
        code_map = builder.get_extra("code_map") or {}
        account_currency_map = builder.get_extra("account_currency_map") or {}
        company_currency_id = builder.get_extra("company_currency_id")
        journal_id = builder.get_journal_id("general")
        company_id = builder._company_id

        move_vals = []
        skipped = 0

        for txn in transactions:
            qbo_id = txn["id"]
            lines_data = txn["lines"]
            if not lines_data:
                skipped += 1
                continue

            first = lines_data[0]
            txn_type = first["type"]
            txn_date = first["date"]
            txn_num = first["num"]
            txn_name = first["name"]

            lines = []
            valid = True

            for ld in lines_data:
                account_code = ld["account_code"]
                account_id = code_map.get(account_code)
                if not account_id:
                    for suffix in [".1", ".2", ".3"]:
                        account_id = code_map.get(account_code + suffix)
                        if account_id:
                            break
                if not account_id:
                    _logger.warning(
                        f"Account {account_code} not found for "
                        f"Journal txn #{qbo_id} ({txn_type})"
                    )
                    valid = False
                    break

                line_name = ld["memo"] or ld["name"] or txn_type
                line_vals = {
                    "account_id": account_id,
                    "debit": ld["debit"],
                    "credit": ld["credit"],
                    "name": line_name,
                }

                # If account has a secondary currency, set it on the
                # line.  Journal amounts are in CAD; for foreign-currency
                # accounts we use the CAD amount as amount_currency too
                # (Odoo will treat it as the foreign amount at a 1:1
                # rate, which is acceptable for historical import where
                # we only need the company-currency TB to match).
                acct_currency = account_currency_map.get(account_id)
                if acct_currency and acct_currency != company_currency_id:
                    foreign_amount = ld["debit"] - ld["credit"]
                    line_vals["currency_id"] = acct_currency
                    line_vals["amount_currency"] = foreign_amount

                lines.append((0, 0, line_vals))

            if not valid or not lines:
                skipped += 1
                continue

            # Verify balance
            total_debit = sum(l[2]["debit"] for l in lines)
            total_credit = sum(l[2]["credit"] for l in lines)
            diff = round(total_debit - total_credit, 2)
            if abs(diff) > 0.01:
                _logger.warning(
                    f"Journal txn #{qbo_id} ({txn_type}) unbalanced "
                    f"by {diff:.2f} — skipping"
                )
                skipped += 1
                continue

            ref = f"JNL-{txn_type}-{qbo_id}"
            narration = (
                f"{txn_type}"
                + (f" #{txn_num}" if txn_num else "")
                + (f" — {txn_name}" if txn_name else "")
            )

            move_vals.append({
                "move_type": "entry",
                "journal_id": journal_id,
                "date": txn_date,
                "ref": ref,
                "narration": narration,
                "company_id": company_id,
                "line_ids": lines,
            })

        _logger.info(
            f"Transformed {len(move_vals)} Journal entries, skipped {skipped}"
        )
        return move_vals

    @ETL.load()
    def load_journal_transactions(
        self, ctx: ETLContext, transformed: Dict
    ) -> None:
        """Load Journal entries into Odoo."""
        move_vals = transformed.get("transform_journal_transactions", [])

        if not move_vals:
            _logger.info("No Journal transactions to create")
            return

        moves = ctx.env["account.move"]
        for vals in move_vals:
            ref = vals.get("ref", "?")
            with ctx.skippable(f"create Journal entry {ref}"):
                moves |= ctx.env["account.move"].create(vals)

        _logger.info(f"Created {len(moves)} Journal entries")

        posted = 0
        by_journal = {}
        for move in moves:
            by_journal.setdefault(move.journal_id.id, self.env["account.move"])
            by_journal[move.journal_id.id] |= move
        for journal_id, journal_moves in sorted(by_journal.items()):
            with post_lock(ctx.env.cr, journal_id):
                for move in journal_moves:
                    with ctx.skippable(
                        f"post Journal entry {move.ref or '?'}"
                    ):
                        move.action_post()
                        posted += 1

        _logger.info(f"Posted {posted} Journal entries")
