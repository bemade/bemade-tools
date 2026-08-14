"""Reusable GL-related helpers for QBO import pipelines.

Functions extracted from ``gl_first_etl.py`` and ``gl_import_etl.py`` for
use by the XLSX fallback pipeline and any other pipeline that needs
GL-level account resolution, XLSX parsing, or journal entry construction
from XLSX export data.
"""

import logging
from collections import defaultdict
from datetime import date, timedelta
from io import BytesIO
from typing import Dict, List, Optional, Set, Tuple

from odoo import _
from odoo.exceptions import UserError
from odoo.fields import Command

_logger = logging.getLogger(__name__)


def build_code_maps(ctx) -> Dict:
    """Build account code, currency, and type maps from Odoo.

    Returns a dict with keys: ``code_map``, ``account_currency_map``,
    ``account_type_map``, ``company_currency_id``, ``currency_name_map``.
    """
    ctx.env.cr.execute(
        """
        SELECT aa.id,
               aa.code_store::jsonb->>jsonb_object_keys(
                   aa.code_store::jsonb) as code,
               aa.currency_id,
               aa.account_type
        FROM account_account aa
        """
    )
    code_map: Dict[str, int] = {}
    account_currency_map: Dict[int, int] = {}
    account_type_map: Dict[int, str] = {}
    for acct_id, code, currency_id, acct_type in ctx.env.cr.fetchall():
        if code:
            code_map[code] = acct_id
            if currency_id:
                account_currency_map[acct_id] = currency_id
            if acct_type:
                account_type_map[acct_id] = acct_type

    ctx.env.cr.execute("SELECT id, name FROM res_currency WHERE active = true")
    currency_name_map = {name: cid for cid, name in ctx.env.cr.fetchall()}

    return {
        "code_map": code_map,
        "account_currency_map": account_currency_map,
        "account_type_map": account_type_map,
        "company_currency_id": ctx.env.company.currency_id.id,
        "currency_name_map": currency_name_map,
    }


def resolve_account(
    code_map: Dict[str, int],
    account_code: str,
    account_type_map: Optional[Dict[int, str]] = None,
) -> Optional[Tuple[int, Optional[str]]]:
    """Resolve account code to (account_id, account_type) or None.

    Tries the code as-is, then with ``.1``, ``.2``, ``.3`` suffixes
    (QBO occasionally exports sub-account variants).
    """
    for suffix in ("", ".1", ".2", ".3"):
        account_id = code_map.get(account_code + suffix)
        if account_id:
            acct_type = account_type_map.get(account_id) if account_type_map else None
            return account_id, acct_type
    return None


def resolve_account_id(code_map: Dict[str, int], account_code: str) -> Optional[int]:
    """Resolve account code to an account_id or None."""
    result = resolve_account(code_map, account_code)
    return result[0] if result else None


def journal_entry_vals_from_export(
    *,
    txn_id: str,
    txn_type: str,
    txn_date: str,
    txn_num: str,
    txn_name: str,
    lines_data: List[Dict],
    code_map: Dict[str, int],
    account_currency_map: Dict[int, int],
    company_currency_id: int,
    txn_currency_id: Optional[int],
    txn_exchange_rate: float,
    journal_id: int,
    company_id: int,
) -> Optional[Dict]:
    """Build a generic journal entry from XLSX GL export lines.

    The QBO Journal export is always in company currency (CAD).  For
    lines on accounts with a foreign currency matching the transaction
    currency, we set ``currency_id`` and ``amount_currency`` using the
    transaction's exchange rate from QBO.
    """
    lines = []
    for ld in lines_data:
        account_code = ld["account_code"]
        account_id = resolve_account_id(code_map, account_code)
        if not account_id:
            _logger.warning(
                "Account %s not found for export txn #%s (%s)",
                account_code, txn_id, txn_type,
            )
            return None

        line_name = ld["memo"] or ld["name"] or txn_type
        line_vals = {
            "account_id": account_id,
            "debit": ld["debit"],
            "credit": ld["credit"],
            "name": line_name,
        }

        acct_currency = account_currency_map.get(account_id)
        if acct_currency and acct_currency != company_currency_id:
            cad_amount = ld["debit"] - ld["credit"]
            if txn_currency_id and txn_currency_id == acct_currency and txn_exchange_rate:
                line_vals["currency_id"] = acct_currency
                line_vals["amount_currency"] = round(cad_amount / txn_exchange_rate, 2)
            else:
                line_vals["currency_id"] = acct_currency
                line_vals["amount_currency"] = cad_amount

        lines.append((0, 0, line_vals))

    if not lines:
        return None

    total_debit = sum(l[2]["debit"] for l in lines)
    total_credit = sum(l[2]["credit"] for l in lines)
    diff = round(total_debit - total_credit, 2)
    if abs(diff) > 0.01:
        _logger.warning(
            "Export txn #%s (%s) unbalanced by %.2f — skipping",
            txn_id, txn_type, diff,
        )
        return None

    ref = f"JNL-{txn_type}-{txn_id}"
    narration = f"{txn_type}" + (f" #{txn_num}" if txn_num else "") + (
        f" — {txn_name}" if txn_name else ""
    )

    return {
        "move_type": "entry",
        "journal_id": journal_id,
        "date": txn_date,
        "ref": ref,
        "narration": narration,
        "company_id": company_id,
        "line_ids": lines,
    }


# ---------------------------------------------------------------------------
# XLSX parsing (from gl_import_etl.py)
# ---------------------------------------------------------------------------

def parse_journal_export(file_content: bytes) -> List[Dict]:
    """Parse a QBO Journal XLSX export into grouped transactions.

    Returns a list of transactions, each with an ``id`` and ``lines``
    (list of dicts with date, type, num, name, memo, account_code,
    account_name, debit, credit).
    """
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
    ws = wb.active

    transactions: List[Dict] = []
    current_id: Optional[str] = None
    current_lines: List[Dict] = []
    has_txn_id_col = False

    for row in ws.iter_rows(min_row=5, max_row=5, values_only=True):
        if len(row) > 10 and row[10] and "Transaction ID" in str(row[10]):
            has_txn_id_col = True
        break

    for row in ws.iter_rows(min_row=6, values_only=True):
        col0 = row[0]
        txn_date = row[1]
        txn_type = row[2]
        txn_num = row[3]
        name = row[4]
        memo = row[5]
        acct_code = row[6]
        acct_name = row[7]
        debit = row[8]
        credit = row[9]
        txn_id_col = row[10] if len(row) > 10 else None

        if col0 and not txn_date:
            col0_str = str(col0).strip()
            if col0_str.startswith("Total for"):
                if current_id and current_lines:
                    transactions.append({"id": current_id, "lines": current_lines})
                current_id = None
                current_lines = []
            elif col0_str.startswith("TOTAL"):
                continue
            else:
                current_id = col0_str
                current_lines = []
            continue

        if txn_type and acct_code is not None:
            if has_txn_id_col and txn_id_col:
                line_txn_id = (
                    str(int(txn_id_col))
                    if isinstance(txn_id_col, float)
                    else str(txn_id_col)
                )
                current_id = line_txn_id

            date_str = str(txn_date) if txn_date else None
            if date_str and "/" in date_str:
                parts = date_str.split("/")
                if len(parts) == 3:
                    date_str = f"{parts[2]}-{parts[1]}-{parts[0]}"

            try:
                d = float(debit) if debit else 0.0
            except (ValueError, TypeError):
                d = 0.0
            try:
                c = float(credit) if credit else 0.0
            except (ValueError, TypeError):
                c = 0.0

            if d == 0 and c == 0:
                continue

            current_lines.append({
                "date": date_str,
                "type": str(txn_type),
                "num": str(txn_num or ""),
                "name": str(name or ""),
                "memo": str(memo or ""),
                "account_code": str(acct_code).strip(),
                "account_name": str(acct_name or ""),
                "debit": d,
                "credit": c,
            })

    wb.close()
    return transactions


# ---------------------------------------------------------------------------
# Imported QBO ID collection (from gl_import_etl.py)
# ---------------------------------------------------------------------------

_QBO_ID_FIELDS = [
    "qbo_invoice_id",
    "qbo_bill_id",
    "qbo_expense_id",
    "qbo_transfer_id",
    "qbo_deposit_id",
    "qbo_journal_entry_id",
    "qbo_credit_memo_id",
    "qbo_vendor_credit_id",
    "qbo_sales_receipt_id",
    "qbo_refund_receipt_id",
    "qbo_tax_payment_id",
    "qbo_cc_payment_id",
    "qbo_payment_id",
    "qbo_bill_payment_id",
]

_QBO_PAYMENT_FIELDS = [
    "qbo_payment_id",
    "qbo_bill_payment_id",
]


def get_imported_qbo_ids(ctx) -> Set[str]:
    """Collect all QBO transaction IDs already in Odoo."""
    imported: Set[str] = set()
    cr = ctx.env.cr

    for field in _QBO_ID_FIELDS:
        cr.execute(
            f"SELECT {field} FROM account_move "
            f"WHERE {field} IS NOT NULL AND {field} != 0"
        )
        imported.update(str(row[0]) for row in cr.fetchall())

    for field in _QBO_PAYMENT_FIELDS:
        cr.execute(
            f"SELECT {field} FROM account_payment "
            f"WHERE {field} IS NOT NULL AND {field} != 0"
        )
        imported.update(str(row[0]) for row in cr.fetchall())

    return imported


# Odoo account_type for Cost-of-Goods-Sold accounts (see account_etl's
# QBO account-type mapping: "Cost of Goods Sold" -> "expense_direct_cost").
_COGS_ACCOUNT_TYPE = "expense_direct_cost"


def build_cogs_completion_transactions(
    ctx,
    cache_id: int,
    imported_ids: Set[str],
    code_map: Dict[str, int],
    account_type_map: Dict[int, str],
) -> List[Dict]:
    """Recover perpetual-COGS lines that entity pipelines structurally drop.

    QBO books perpetual COGS (DR Cost of Goods Sold / CR inventory
    valuation) directly on inventory-item sales.  The entity pipelines
    import only the API ``Line`` array (revenue / AR / tax), and the
    invoice poster deliberately skips Odoo's own COGS generation — which
    cannot fire anyway on link-less imported invoices (no sale-order or
    delivery to relieve).  ``journal_fallback`` then excludes any
    transaction an entity pipeline imported, so those COGS / inventory
    lines are recovered nowhere and inventory is never relieved.

    For each cached transaction that WAS entity-imported, return the COGS /
    inventory lines the Odoo move never booked — scoped to a Cost-of-Goods-
    Sold account or QBO's inventory account, and only when:

    * the scoped missing lines self-balance (a clean COGS pair balances
      exactly; this drops any unrelated FX/rounding line QBO booked on the
      revenue side, which would otherwise unbalance the completion), and
    * at least one scoped line lands on a Cost-of-Goods-Sold account
      (so a coincidental balancing subset on other accounts is never
      plugged).

    QBO's inventory account is discovered from its own "Inventory Starting
    Value" transactions (the account they debit), NOT from Odoo's category
    valuation config — the migration leaves Odoo's valuation accounts
    Odoo-correct for the future, but QBO historically relieved a different
    inventory account, which is where the opening balance actually sits.

    Returns transactions shaped like ``get_transactions_for_import`` so the
    fallback transform builds balancing JEs from them unchanged.
    """
    if not imported_ids:
        return []
    cr = ctx.env.cr

    # account_id -> code, from the (company-agnostic) code_map, so the
    # booked-codes query doesn't assume a particular company key.
    id_to_code = {aid: code for code, aid in code_map.items()}

    # QBO's inventory account(s): those debited in its own Inventory Starting
    # Value transactions.  Odoo's category valuation accounts differ and are
    # not where QBO booked the opening balance or the COGS relief.
    cr.execute(
        """
        SELECT l.account_code
        FROM qbo_journal_cache_transaction t
        JOIN qbo_journal_cache_line l ON l.transaction_id = t.id
        WHERE t.cache_id = %s AND t.txn_type = 'Inventory Starting Value'
        GROUP BY l.account_code
        HAVING sum(l.debit) > sum(l.credit)
        """,
        (cache_id,),
    )
    inventory_codes = {row[0] for row in cr.fetchall() if row[0]}

    def _in_cogs_scope(code: str) -> bool:
        return (
            account_type_map.get(code_map.get(code or "")) == _COGS_ACCOUNT_TYPE
            or (code or "") in inventory_codes
        )

    # 1. Account codes each entity-imported move actually booked, keyed by
    #    QBO transaction id (pooled across every move-type id field).
    booked: Dict[str, set] = defaultdict(set)
    for field in _QBO_ID_FIELDS:
        cr.execute(
            f"""
            SELECT am.{field}, aml.account_id
            FROM account_move am
            JOIN account_move_line aml ON aml.move_id = am.id
            WHERE am.{field} IS NOT NULL AND am.{field} != 0
            """
        )
        for qbo_id, account_id in cr.fetchall():
            code = id_to_code.get(account_id)
            if code:
                booked[str(qbo_id)].add(code)

    # 2. Cache lines for the entity-imported transactions.
    cr.execute(
        """
        SELECT t.qbo_txn_id, t.txn_type, t.txn_date, t.txn_num, t.txn_name,
               l.account_code, l.account_name, l.memo, l.name,
               l.debit, l.credit
        FROM qbo_journal_cache_transaction t
        JOIN qbo_journal_cache_line l ON l.transaction_id = t.id
        WHERE t.cache_id = %s AND t.qbo_txn_id = ANY(%s)
        ORDER BY t.qbo_txn_id
        """,
        (cache_id, list(imported_ids)),
    )
    txns: Dict[str, Dict] = {}
    for (qbo_id, ttype, tdate, tnum, tname,
         acode, aname, memo, lname, debit, credit) in cr.fetchall():
        txn = txns.setdefault(qbo_id, {
            "type": ttype, "date": tdate, "num": tnum,
            "name": tname, "lines": [],
        })
        txn["lines"].append({
            "account_code": acode, "account_name": aname,
            "memo": memo, "name": lname,
            "debit": debit or 0.0, "credit": credit or 0.0,
        })

    result: List[Dict] = []
    for qbo_id, txn in txns.items():
        booked_codes = booked.get(qbo_id, set())
        missing = [
            ln for ln in txn["lines"]
            if (ln["account_code"] or "") not in booked_codes
            and _in_cogs_scope(ln["account_code"])
        ]
        if not missing:
            continue
        diff = round(
            sum(l["debit"] for l in missing)
            - sum(l["credit"] for l in missing), 2
        )
        if abs(diff) > 0.01:
            continue  # not a clean self-balancing pair — never plug
        has_cogs = any(
            account_type_map.get(code_map.get(l["account_code"] or ""))
            == _COGS_ACCOUNT_TYPE
            for l in missing
        )
        if not has_cogs:
            continue
        result.append({
            "id": qbo_id,
            "lines": [{
                "date": str(txn["date"]) if txn["date"] else "",
                "type": txn["type"] or "",
                "num": txn["num"] or "",
                "name": txn["name"] or "",
                "memo": l["memo"] or "",
                "account_code": l["account_code"] or "",
                "account_name": l["account_name"] or "",
                "debit": l["debit"],
                "credit": l["credit"],
            } for l in missing],
        })
    return result


# ---------------------------------------------------------------------------
# Year-end closing entries (qbo.year.end.close)
# ---------------------------------------------------------------------------

# ir.config_parameter keys gating and parameterizing the pass. Read at load
# time so a migration that hasn't opted in never touches the ledger.
YEAR_END_CLOSE_ENABLED_PARAM = "qbo_to_odoo.year_end_close.enabled"
YEAR_END_CLOSE_RE_CODE_PARAM = "qbo_to_odoo.year_end_close.retained_earnings_code"
DEFAULT_RETAINED_EARNINGS_CODE = "3200"

# account.account.internal_group values that make up P&L (covers every
# income_*/expense_* account_type, since internal_group is that type's
# leading segment).
_PNL_INTERNAL_GROUPS = ("income", "expense")

# Idempotency / drift-exclusion stamp, mirroring the QBO_EXCH:/QBO_CENT_TRUEUP
# ref conventions used elsewhere in this module.
YEAR_END_CLOSE_REF_PREFIX = "QBO_YE_CLOSE:"


def get_year_end_close_config(env) -> Dict:
    """Read the year-end-close gate and destination-account config.

    :return: ``{"enabled": bool, "retained_earnings_code": str}`` —
        disabled and ``3200`` by default, so a migration that hasn't
        opted in via ``ir.config_parameter`` never posts closing entries.
    """
    icp = env["ir.config_parameter"].sudo()
    return {
        "enabled": icp.get_param(YEAR_END_CLOSE_ENABLED_PARAM, "0") == "1",
        "retained_earnings_code": icp.get_param(
            YEAR_END_CLOSE_RE_CODE_PARAM, DEFAULT_RETAINED_EARNINGS_CODE
        ),
    }


def get_unaffected_earnings_account(env, company):
    """Resolve *company*'s Current-Year-Earnings account.

    Searches ``active_test=False`` for ``account_type == 'equity_unaffected'``
    scoped to *company* — deliberately NOT
    ``company.get_unaffected_earnings_account()``, which searches active
    accounts only and, finding none because the real one is archived
    (322000 in this migration, archived because QBO marked it inactive and
    ``qbo.account.finalizer`` mirrors that), would *create a duplicate
    999999 account* (see ``account/models/company.py`` around
    ``get_unaffected_earnings_account``).

    :raises UserError: if more than one CYE account resolves — the pass
        must never split the year-end-close allocation across two
        accounts; a stray active duplicate has to be resolved by hand.
    :return: the ``account.account`` recordset (empty if none found).
    """
    Account = env["account.account"].with_context(active_test=False)
    accounts = Account.search([
        *Account._check_company_domain(company),
        ("account_type", "=", "equity_unaffected"),
    ])
    if len(accounts) > 1:
        raise UserError(_(
            "Multiple Current-Year-Earnings accounts found for %(company)s: "
            "%(codes)s. Year-end close cannot proceed until only one "
            "remains.",
            company=company.display_name,
            codes=", ".join(accounts.mapped("code")),
        ))
    return accounts


def get_retained_earnings_account(env, company, config: Dict):
    """Resolve *company*'s configured Retained Earnings account by code.

    The code comes from *config* (see ``get_year_end_close_config``,
    default ``3200``), so a chart that doesn't use that literal can be
    repointed via the ``qbo_to_odoo.year_end_close.retained_earnings_code``
    ``ir.config_parameter`` without a code change. Searched
    ``active_test=False`` for symmetry with the CYE lookup; unlike the CYE
    account, an archived RE account is not auto-unarchived here — 322000
    being archived is the specific, confirmed trap this design guards
    against, not the RE account.

    :return: the ``account.account`` recordset (empty if the code doesn't
        resolve).
    """
    Account = env["account.account"].with_context(active_test=False)
    return Account.search([
        *Account._check_company_domain(company),
        ("code", "=", config["retained_earnings_code"]),
    ], limit=1)


def validate_year_end_close_accounts(env, company, config: Dict):
    """Resolve and validate the CYE/RE accounts, failing loudly on
    misconfiguration.

    Deliberately only ever called from the *load* step, after the
    ``enabled`` gate has already been checked — a migration that hasn't
    opted into year-end close must never be affected by these checks,
    even if its chart of accounts happens to have an unrelated account
    coded like the default ``3200`` or two accounts typed
    ``equity_unaffected``.

    :raises UserError: if the Current-Year-Earnings account does not
        resolve to exactly one record, if the Retained Earnings account
        does not resolve to exactly one record, or if the resolved RE
        account is not a carry-forward equity account
        (``account_type == 'equity'``) — pointing it at a P&L or
        ``equity_unaffected`` account would leave the report's
        Undistributed Profits/Losses line non-zero instead of clearing it
        (see ``account_account._compute_include_initial_balance``).
    :return: ``(cye_account, re_account)`` singleton recordsets.
    """
    cye_account = get_unaffected_earnings_account(env, company)
    if not cye_account:
        raise UserError(_(
            "Year-end close is enabled but no Current-Year-Earnings "
            "account (account_type = 'equity_unaffected') was found for "
            "%(company)s.",
            company=company.display_name,
        ))

    re_account = get_retained_earnings_account(env, company, config)
    if not re_account:
        raise UserError(_(
            "Year-end close is enabled but the configured Retained "
            "Earnings account (code %(code)s) was not found for "
            "%(company)s.",
            code=config["retained_earnings_code"],
            company=company.display_name,
        ))
    if re_account.account_type != "equity":
        raise UserError(_(
            "Year-end close: the configured Retained Earnings account "
            "%(code)s must be a carry-forward equity account "
            "(account_type = 'equity'); found %(type)s instead. Posting "
            "would leave the Undistributed Profits/Losses line non-zero.",
            code=re_account.code,
            type=re_account.account_type,
        ))

    return cye_account, re_account


def get_year_end_close_journal(env, company):
    """Resolve the miscellaneous journal for the closing entries.

    Same convention as ``qbo.migration.report``'s opening-balance entry:
    prefer the journal coded ``MISC``, fall back to any ``general``
    journal for *company*.
    """
    Journal = env["account.journal"]
    domain = [("type", "=", "general"), ("company_id", "=", company.id)]
    journal = Journal.search(domain + [("code", "=", "MISC")], limit=1)
    if not journal:
        journal = Journal.search(domain, limit=1)
    return journal


def compute_fiscal_year_close_rows(
    env, company, config: Dict
) -> List[Tuple[int, date, float, int, int]]:
    """Compute one closing row per already-closed fiscal year.

    Enumerates fiscal years from the one containing the earliest posted
    P&L move line up to (but excluding) the fiscal year containing today,
    with boundaries always read from ``company.compute_fiscalyear_dates``
    (never hardcoded — a change to ``fiscalyear_last_month``/``_day``
    shifts every result). For each FY, sums the ``balance`` (debit −
    credit) of every posted move line on an income/expense account dated
    within that FY's window; a negative sum is a net profit.

    Side-effect-free (only reads), so it is testable directly against a
    seeded DB without going through the ``qbo.year.end.close`` pipeline —
    the same seam ``build_cogs_completion_transactions`` gives the COGS
    completion logic.

    :param env: Odoo environment.
    :param company: a single ``res.company`` record.
    :param config: as returned by ``get_year_end_close_config`` — only
        ``retained_earnings_code`` is used here.
    :return: a list of ``(fy_label, date_to, signed_result,
        cye_account_id, re_account_id)`` tuples, one per closed FY whose
        signed result is non-zero (a zero-result FY needs no closing
        entry and is omitted). ``fy_label`` is ``date_to.year``.
    """
    cye_account = get_unaffected_earnings_account(env, company)
    re_account = get_retained_earnings_account(env, company, config)
    if not cye_account or not re_account:
        _logger.warning(
            "Year-end close: could not resolve the Current-Year-Earnings "
            "and/or Retained Earnings account for %s — nothing to close",
            company.display_name,
        )
        return []

    # active_test=False is load-bearing: Odoo's report-only "Undistributed
    # Profits/Losses" line resolves its accounts by account_type and never
    # filters on `active`, so any archived P&L account skipped here becomes
    # permanent drift between the posted Retained Earnings balance and the
    # reported one. Migrated books make that the common case — the QBO
    # account pipeline archives every account the source chart no longer
    # uses, and those routinely carry years of historical activity.
    pnl_accounts = env["account.account"].with_context(active_test=False).search([
        *env["account.account"]._check_company_domain(company),
        ("internal_group", "in", list(_PNL_INTERNAL_GROUPS)),
    ])
    if not pnl_accounts:
        return []

    AccountMoveLine = env["account.move.line"]
    base_domain = [
        ("parent_state", "=", "posted"),
        ("account_id", "in", pnl_accounts.ids),
        ("company_id", "=", company.id),
    ]

    earliest = AccountMoveLine.search(base_domain, order="date asc", limit=1)
    if not earliest:
        return []

    open_fy = company.compute_fiscalyear_dates(date.today())
    fy_window = company.compute_fiscalyear_dates(earliest.date)

    rows: List[Tuple[int, date, float, int, int]] = []
    while fy_window["date_from"] < open_fy["date_from"]:
        date_from, date_to = fy_window["date_from"], fy_window["date_to"]
        lines = AccountMoveLine.search(
            base_domain + [("date", ">=", date_from), ("date", "<=", date_to)]
        )
        signed_result = round(sum(lines.mapped("balance")), 2)
        fy_label = date_to.year

        if signed_result:
            rows.append(
                (fy_label, date_to, signed_result, cye_account.id, re_account.id)
            )

        fy_window = company.compute_fiscalyear_dates(date_to + timedelta(days=1))

    return rows


def build_year_end_close_move_vals(
    fy_label: int,
    date_to: date,
    signed_result: float,
    cye_account_id: int,
    re_account_id: int,
    journal_id: int,
) -> Dict:
    """Build a balanced year-end closing ``account.move`` vals dict.

    Pure / side-effect-free — no DB access, just arithmetic and a vals
    dict — unit-testable without a database, mirroring
    ``build_cogs_completion_transactions``'s DB-free construction seam.

    Sign convention: *signed_result* is the FY's posted P&L balance
    (debit − credit); negative means a net profit. The Retained Earnings
    leg's balance changes by ``+signed_result`` and the Current-Year-
    Earnings leg's balance changes by ``-signed_result`` — so a profit
    year (negative ``signed_result``) debits the CYE account and credits
    Retained Earnings by the profit amount (CYE clears toward zero, RE
    grows); a loss year reverses both legs.

    :return: an ``account.move`` vals dict, balanced by construction,
        dated ``date_to``, with ``ref = f"{YEAR_END_CLOSE_REF_PREFIX}
        {fy_label}"`` — the idempotency / drift-exclusion stamp.
    """
    return {
        "move_type": "entry",
        "journal_id": journal_id,
        "date": date_to,
        "ref": f"{YEAR_END_CLOSE_REF_PREFIX}{fy_label}",
        "line_ids": [
            Command.create({
                "account_id": cye_account_id,
                "name": f"Year-end close FY{fy_label}",
                "debit": max(-signed_result, 0.0),
                "credit": max(signed_result, 0.0),
            }),
            Command.create({
                "account_id": re_account_id,
                "name": f"Year-end close FY{fy_label}",
                "debit": max(signed_result, 0.0),
                "credit": max(-signed_result, 0.0),
            }),
        ],
    }
